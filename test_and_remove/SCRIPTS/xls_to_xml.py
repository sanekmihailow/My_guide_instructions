from openpyxl import load_workbook
import datetime
import xml.etree.cElementTree as ET
import pprint
from yattag import Doc, indent

wb = load_workbook("Learn.xlsx")
ws = wb.worksheets[0]
# ls = dir(ws)
max_rows = ws._current_row
# pprint.pprint(ws._cells)
doc, tag, text = Doc().tagtext()
xml_header = '<?xml version="1.0" encoding="UTF-8"?>'
xml_schema = '<RegistrySet xsi:noNamespaceSchemaLocation="schema.xsd" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
doc.asis(xml_header)
doc.asis(xml_schema)

for row in ws.iter_rows(min_row=2, max_row=max_rows, min_col=1, max_col=14):
    with tag("RegistryRecord"):
        # print([cell.value for cell in row])
        row = [cell.value for cell in row]
        with tag("Worker"):
            with tag("LastName"):
                text(row[0])
            with tag("FirstName"):
                text(row[1])
            with tag("MiddleName"):
                text(row[2])
            with tag("Snils"):
                text(row[3])
            with tag("Position"):
                text(row[4])
            with tag("EmployerInn"):
                text(row[5])
            with tag("EmployerTitle"):
                text(row[6])

        with tag("Organization"):
            with tag("Inn"):
                text(row[7])
            with tag("Title"):
                text(row[8])

        with tag("Test", isPassed=row[9], learnProgramId=row[10]):
            #            with tag("isPassed"):
            #                text(row[9])
            #            with tag("learnProgramId"):
            #                text(row[10])
            with tag("Date"):
                text(row[11])
            with tag("ProtocolNumber"):
                text(row[12])
            with tag("LearnProgramTitle"):
                text(row[13])


result = indent(
    doc.getvalue(),
    indentation='  ',
    indent_text=False
)
with open("output.xml", "w", encoding="utf-8") as f:
    f.write(result)
    lasttag = "\n</RegistrySet>"
    f.write(lasttag)
# pprint.pprint(ws._cells)
"""
3.11
pip3 install yattag pprit autopep8 openpyxl
"""
